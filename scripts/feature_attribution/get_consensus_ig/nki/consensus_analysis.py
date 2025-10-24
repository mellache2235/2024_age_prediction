# Implementation based on:
#   https://github.com/scsnl/YZ_HCP_DNN_Gender_2023/blob/main/dnn/get_consensus_features_multiple_cv.py
#   https://github.com/scsnl/YZ_HCP_DNN_Gender_2023/blob/4a25f837152ce6cb5dbeb226521f40944e8c1b1e/dnn/utilityFunctions.py#L87
#
# Assumes IG-based feature attributions have been extracted for multiple models and stored as npz files at
# $OAK/projects/ansharma/2023_AS_MLD_DNN/results/restfmri/dnn/featureattributions/{TRAIN_SET}
#
# Finds consensus features by getting the top xth percentile features and counting occurrences across models
# Adapted from Yuan Zhang's scripts for HCP sex classification by Anirudh Sharma, 06/04/2024.
# Updated 06/07/2024 for regression.

import numpy as np
import os
import collections
import json
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl import Workbook, load_workbook
import math
import nilearn
from nilearn import image, plotting
from collections import Counter
import sys

sys.path.append('/oak/stanford/groups/menon/projects/mellache/2024_age_prediction/scripts/train_regression_models/')
from utility_functions import *

REGRESSION = True

TRAIN_SET = "hcp_dev"
# TRAIN_SET = "scsnl_tdmd_2024"
# SUB_DIR = ""
SUB_DIR = "06_26_2024"
ig_dir = '/oak/stanford/groups/menon/projects/mellache/2024_age_prediction/results/figures/nki/ig_files/'
num_models = 500
# num_models = 5 * 2
# num_models = 2
#percentiles = [90]
percentiles = [50,80]
groups = {"TD": 0}
if REGRESSION:
    groups = {"continuous": 999}
bn_atlas_file = "/oak/stanford/groups/menon/projects/sryali/2019_DNN/scripts/features/bnatlas_tree.json"
with open(bn_atlas_file) as f:
    bn_atlas = json.load(f)
parent_substitutions = {
    "Amyg, Amygdala": "Middle Temporal Gyrus",
    "CG, Cingulate Gyrus": "Cingulate Gyrus",
    "Cun, Cuneus": "PCC, Precuneus",
    "Frontal Lobe": "Prefrontal Cortex",
    "FuG, Fusiform Gyrus": "Inferior Temporal Gyrus",
    "Hipp, Hippocampus": "Middle Temporal Gyrus",
    "IFG, Inferior Frontal Gyrus": "Prefrontal Cortex",
    "INS, Insular Gyrus": "Prefrontal Cortex",
    "IPL, Inferior Parietal Lobule": "Inferior Parietal Lobe",
    "ITG, Inferior Temporal Gyrus": "Inferior Temporal Gyrus",
    "Insular Lobe": "Prefrontal Cortex",
    "Limbic Lobe": "Middle Temporal Lobe",
    "MFG, Middle Frontal Gyrus": "Prefrontal Cortex",
    "MTG, Middle Temporal Gyrus": "Middle Temporal Gyrus",
    "OcG, Occipital Gyrus": "Occipital Gyrus",
    "Occipital Lobe": "Occipital Lobe",
    "OrG, Orbital Gyrus": "Prefrontal Cortex",
    "PCL,Paracentral Lobule": "Paracentral Lobule",
    "Parietal Lobe": "Parietal Lobe",
    "Pcun, Precuneus": "PCC, Precuneus",
    "PhG, Parahippocampal Gyrus": "Middle Temporal Gyrus",
    "PoG, Postcentral Gyrus": "Postcentral Gyrus",
    "PrG, Precentral Gyrus": "Precentral Gyrus",
    "Psts, Posterior Superior Temporal Sulcus": "Superior Temporal Gyrus",
    "SFG, Superior Frontal Gyrus": "Prefrontal Cortex",
    "SPL, Superior Parietal Lobule": "Superior Parietal Lobule",
    "STG, Superior Temporal Gyrus": "Superior Temporal Gyrus",
    "Str, Striatum": "Striatum",
    "Subcortical Nuclei": "Subcortical Nuclei",
    "Temporal Lobe": "Temporal Lobe",
    "Tha, Thalamus": "Thalamus",
}

def save_feature_consensus_nifti(occurrences, group, percentile):
    bn_nifti = '/oak/stanford/groups/menon/projects/sryali/2019_DNN/scripts/features/BN_Atlas_246_2mm.nii'

    atlas_volume = image.load_img(bn_nifti)
    roi_nifti = image.math_img('img-img', img=atlas_volume)
    img_data = atlas_volume.get_data()

    for feature in occurrences:
        # print(feature, type(feature), occurrences[feature], type(occurrences[feature]))
        roi_idx = np.where(img_data == feature + 1, (occurrences[feature]*1.0), 0)
        roi_img = image.new_img_like(roi_nifti, roi_idx)
        roi_nifti = image.math_img('img1+img2', img1=roi_nifti, img2=roi_img)

    output_nii_file = 'bn_features_consensus_group_%s_percentile_%02d.nii.gz'%(group, percentile)
    print(output_nii_file)
    roi_nifti.to_filename(output_nii_file)

path = "/oak/stanford/groups/menon/projects/mellache/2021_foundation_model/data/imaging/for_dnn/nki_age_cog_dev/fold_0.bin"
X_train, X_valid, Y_train, Y_valid = load_finetune_dataset(path)
X_train = reshapeData(X_train)
X_valid = reshapeData(X_valid)
X_total = np.concatenate((X_train,X_valid))
Y_total = np.concatenate((Y_train,Y_valid))

for percentile in percentiles:
    print(f"Percentile={percentile}")
    for group in groups.keys():
        print(f"Group={group}")
        # get top features of each fold and count
        # the number of appearance of each feature across folds
        # and write features to excel
        feature_data = []
        for k in range(num_models):
            print(f"Model_idx={k}")
            ig_k = np.load(
                os.path.join(ig_dir, f"dev_age_model_{k}_ig_nki_cog_dev.npz"),
                allow_pickle=True,
            )
            attr_data = ig_k["arr_0"]  # no_subjs x no_channels x no_ts
            
            if not REGRESSION:
                labels = Y_total
            #preds = ig_k["preds"]
            #pid = ig_k["pid"]

            # print(attr_data.shape)
            # print(labels.shape)
            # print(labels == groups[group])
            if REGRESSION:
                group_features = attr_data
            else:
                group_features = attr_data[labels == groups[group]]
            medians = np.median(group_features, axis=2)  # medians across time points
            mean_across_subj = np.mean(
                np.abs(medians), axis=0
            )  # average abs medians across subjects
            percentiles = np.where(
                np.abs(mean_across_subj)
                >= np.percentile(np.abs(mean_across_subj), percentile)
            )
            feature_idcs = percentiles[
                0
            ]  # includes all indices (rois) at which position the values are above the cutoff
            features = mean_across_subj  # feature scores (averaged across subjects)

            feature_data.append(feature_idcs)

        print(len(feature_data))
        # get all top ROI/feature IDs from all folds and flatten all lists into one list
        feature_data_flatten = [ft for sublist in feature_data for ft in sublist]
        print(len(feature_data_flatten))
        # count the occurrence of each top feature
        occurrences = collections.Counter(feature_data_flatten)
        #save_feature_consensus_nifti(occurrences, group, percentile)
        print("Occurrences of each top feature:")
        print(occurrences)
        
        # occurrences is a dictionary with keys = ROI/feature IDs, and values = #occurrence
        # write results to excel
        thresh = 100 - percentile
        excel_file = os.path.join(ig_dir, f"top_{thresh}_consensus_features_nki_cog_dev_aging.xlsx")
        if not os.path.exists(excel_file):
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "%s_%02d" % (group, percentile)
        else:
            wb = load_workbook(excel_file)
            ws1 = wb.create_sheet(title="%s_%02d" % (group, percentile))
        ws1.append(
            [
                "Region ID",
                "Gyrus",
                "Description",
                "Region Alias",
                "(ID) Region Label",
                "Count",
            ]
        )
        feature_data = []
        for feature in occurrences:
            # print(feature, type(feature), occurrences[feature], type(occurrences[feature]))
            featureID = str(feature + 1)  # ROI/feature ID
            for idx, region in enumerate(bn_atlas):
                if region["id"] == featureID:
                    # print("feature: {}, frequency: {}\n".format(featureID, occurrences[feature]))
                    feature_data.append(
                        [
                            featureID,
                            parent_substitutions[region["parent"]],
                            region["text"],
                            region["data"]["alias"],
                            "(%s), %s" % (featureID, region["text"]),
                            "%01d" % (occurrences[feature]),
                        ]
                    )
        for feature in feature_data:
            ws1.append(feature)
        tab = Table(
            displayName="group%s_percentile%02d" % (group, percentile),
            ref="A1:F%d" % (len(occurrences) + 1),
        )
        style = TableStyleInfo(
            name="TableStyleLight15",
            showFirstColumn=False,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws1.add_table(tab)
        table_cells = ws1["A1" : "F%d" % (len(occurrences) + 1)]
        cols = ["A", "B", "C", "D", "E", "F"]
        for idx, col in enumerate(ws1.columns):
            max_length = 0
            column = col[0].column
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws1.column_dimensions[cols[idx]].width = math.ceil(adjusted_width)
        for row in table_cells:
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrapText=True
                )
        print("Saving to Excel...")
        wb.save(filename=excel_file)
        # # write results to nii for plot
        # save_feature_consensus_nifti(
        #     occurrences, output_nii_path, group, site, percentile
        # )
